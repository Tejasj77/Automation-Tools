import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

name_of_file = input("Enter the path")
main_frame = pd.read_excel(f'{name_of_file}',keep_default_na = False)
signal_name = list(main_frame["Signal Name"])



def dataframe_printing(df,name):
    df.to_excel(f'{name}.xlsx',index=False)
    """
    book = xw.Book()
    shit = book.sheets['Sheet1']
    shit.range('A1').value = df
    :param df: 
    :return: 
    """

dataframe_list = []
for i in range(0,len(main_frame)):
    dataframe_list.append(main_frame[i:i+1])

for i in range(len(dataframe_list)):
    dataframe_printing(dataframe_list[i],signal_name[i])
    wb = load_workbook(filename=f'{signal_name[i]}.xlsx')
    ws = wb.active
    ws.column_dimensions['A'].width = 27
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 19
    ws.column_dimensions['J'].width = 19
    ws.column_dimensions['K'].width = 19
    max_column = ws.max_column
    for x in range(2,max_column+1):
        col = ws.cell(column=x,row=2)
        col.alignment = Alignment(horizontal='center',wrap_text=True)
        print(col.value)
    wb.save(f'{signal_name[i]}.xlsx')

